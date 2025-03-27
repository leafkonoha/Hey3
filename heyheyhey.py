import requests
import threading
import openpyxl
import urllib3
from openpyxl.styles import PatternFill

# Suppress SSL warnings for unverified HTTPS requests
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ✅ Function to Check HP iLO Health Status
def check_hp_ilo_health(ip, username, password):
    url = f"https://{ip}/redfish/v1/Systems/1"  # HP iLO Redfish API
    try:
        response = requests.get(url, auth=(username, password), verify=False, timeout=5)
        response.raise_for_status()
        data = response.json()

        health_status = data.get("Status", {}).get("Health", "Unknown")  # Extract health status

        # ✅ Extract failed components if health is NOT OK
        failed_components = []
        if health_status.lower() != "ok":
            components = data.get("Oem", {}).get("Hpe", {}).get("AggregateHealthStatus", {})
            for key, value in components.items():
                if value.lower() != "ok":
                    failed_components.append(f"{key}: {value}")

        failure_details = ", ".join(failed_components) if failed_components else "None"
        print(f"[HP iLO] {ip} - Health: {health_status} - Issues: {failure_details}")
        return health_status, failure_details

    except requests.exceptions.RequestException as e:
        print(f"[HP iLO] {ip} - ERROR: {str(e)}")
        return "Error", str(e)


# ✅ Function to Check Dell iDRAC (Redfish) Health Status
def check_dell_redfish_health(ip, username, password):
    url = f"https://{ip}/redfish/v1/Systems/System.Embedded.1"
    try:
        response = requests.get(url, auth=(username, password), verify=False, timeout=5)
        response.raise_for_status()
        data = response.json()

        health_status = data.get("Status", {}).get("Health", "Unknown")

        # ✅ Extract failed components if health is NOT OK
        failed_components = []
        components = ["Power", "Processors", "Memory", "Storage", "NetworkAdapters"]
        if health_status.lower() != "ok":
            for component in components:
                component_status = data.get(component, {}).get("Status", {}).get("Health", "Unknown")
                if component_status.lower() != "ok":
                    failed_components.append(f"{component}: {component_status}")

        failure_details = ", ".join(failed_components) if failed_components else "None"
        print(f"[Dell iDRAC] {ip} - Health: {health_status} - Issues: {failure_details}")
        return health_status, failure_details

    except requests.exceptions.RequestException as e:
        print(f"[Dell iDRAC] {ip} - ERROR: {str(e)}")
        return "Error", str(e)


# ✅ Function to Determine Server Type & Route to Correct Health Check
def check_server_health(ip, username, password, results):
    if ip.lower().startswith("uname"):  # HP iLO
        health_status, failure_details = check_hp_ilo_health(ip, username, password)
        results.append([ip, health_status, failure_details])
    elif ip.lower().startswith("bltwa"):  # Dell iDRAC (Redfish)
        health_status, failure_details = check_dell_redfish_health(ip, username, password)
        results.append([ip, health_status, failure_details])
    else:
        print(f"[SKIPPED] {ip} - Not an HP or Dell server")


# ✅ Function to Save Results in an Excel File (With Color Coding)
def save_to_xlsx(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Server Health Report"

    # ✅ Headers
    headers = ["Server", "Health Status", "Issues"]
    ws.append(headers)

    # ✅ Define Color Coding
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for OK
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for Warning
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for Error

    for row in results:
        ws.append(row)
        cell = ws.cell(row=ws.max_row, column=2)

        # Apply color coding
        if row[1].lower() == "ok":
            cell.fill = green_fill
        elif row[1].lower() == "warning":
            cell.fill = yellow_fill
        elif "error" in row[1].lower():
            cell.fill = red_fill

    # ✅ Save XLSX file
    wb.save("server_health_report.xlsx")
    print("✅ Report saved: server_health_report.xlsx")


# ✅ Main Function
def main():
    # 📝 Servers List (Replace with File Reading If Needed)
    servers = [
        "uname-server1.example.com",  # HP iLO
        "bltwa-server2.example.com",  # Dell iDRAC
        "random-server3.example.com"  # This will be ignored
    ]
    username = "admin"
    password = "password"

    results = []
    threads = []

    for ip in servers:
        thread = threading.Thread(target=check_server_health, args=(ip, username, password, results))
        thread.start()
        threads.append(thread)

    for thread in threads:
        thread.join()

    # ✅ Save results to XLSX
    save_to_xlsx(results)

    print("✅ Health check completed. Results saved in 'server_health_report.xlsx'.")


# ✅ Run the script
if __name__ == "__main__":
    main()
