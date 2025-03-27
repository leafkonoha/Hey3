import requests
import redfish
import hpilo
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Read credentials from a separate file
def read_credentials():
    with open("credentials.txt", "r") as file:
        creds = dict(line.strip().split("=") for line in file)
    return creds["username"], creds["password"]

# Read servers from a cluster-separated text file
def read_servers(file_path):
    clusters = {}
    current_cluster = None

    with open(file_path, "r") as file:
        for line in file:
            line = line.strip()
            if line.startswith("Cluster:"):
                current_cluster = line.replace("Cluster:", "").strip()
                clusters[current_cluster] = []
            elif line and current_cluster:
                clusters[current_cluster].append(line)

    return clusters

# Function to detect the server type
def detect_server_type(hostname, username, password):
    try:
        ilo = hpilo.Ilo(hostname, login=username, password=password, timeout=5)
        ilo.get_fw_version()
        return "HP iLO"
    except:
        pass

    try:
        url = f"https://{hostname}/redfish/v1/Systems/System.Embedded.1"
        response = requests.get(url, auth=(username, password), verify=False, timeout=5)
        if response.status_code == 200:
            return "Dell Redfish"
    except:
        pass

    return "Unknown"

# Fetch health data from HP iLO
def fetch_hp_ilo_health(hostname, username, password):
    try:
        ilo = hpilo.Ilo(hostname, login=username, password=password, timeout=5)
        health_data = ilo.get_embedded_health()
        return {
            "hostname": hostname,
            "server_type": "HP iLO",
            "health_status": health_data['summary']['health'],
        }
    except Exception as e:
        return {"hostname": hostname, "server_type": "HP iLO", "error": str(e)}

# Fetch health data from Dell Redfish
def fetch_dell_redfish_health(hostname, username, password):
    try:
        url = f"https://{hostname}/redfish/v1/Systems/System.Embedded.1"
        session = redfish.redfish_client(base_url=f"https://{hostname}", username=username, password=password, timeout=5)
        session.login(auth="session")
        response = session.get(url)
        session.logout()

        if response.status == 200:
            data = response.dict
            return {
                "hostname": hostname,
                "server_type": "Dell Redfish",
                "health_status": data.get("Status", {}).get("Health", "Unknown"),
            }
    except Exception as e:
        return {"hostname": hostname, "server_type": "Dell Redfish", "error": str(e)}

# Process a single server
def process_server(cluster_name, hostname, username, password):
    server_type = detect_server_type(hostname, username, password)
    
    if server_type == "HP iLO":
        result = fetch_hp_ilo_health(hostname, username, password)
    elif server_type == "Dell Redfish":
        result = fetch_dell_redfish_health(hostname, username, password)
    else:
        result = {"hostname": hostname, "server_type": "Unknown", "error": "Could not detect server type"}

    result["cluster"] = cluster_name  # Include cluster name
    return result

# Save data to an Excel file with clusters and color coding
def save_to_excel(results, output_file):
    df = pd.DataFrame(results)
    df.to_excel(output_file, index=False)

    wb = Workbook()
    ws = wb.active

    # Headers
    headers = df.columns.tolist()
    ws.append(headers)

    # Color codes
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for OK
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red for error
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow for Warning

    # Keep track of cluster rows
    last_cluster = None

    for index, row in df.iterrows():
        row_values = row.tolist()

        # If cluster changes, insert a row
        if last_cluster != row["cluster"]:
            ws.append([f"Cluster: {row['cluster']}"])
            last_cluster = row["cluster"]

        ws.append(row_values)

        # Color health_status
        health_status = row["health_status"] if "health_status" in row else ""
        cell = ws.cell(row=ws.max_row, column=headers.index("health_status") + 1)

        if health_status == "OK":
            cell.fill = green_fill
        elif health_status == "Warning":
            cell.fill = yellow_fill
        elif health_status in ["Critical", "Failed"] or "error" in row:
            cell.fill = red_fill

    wb.save(output_file)
    print(f"Health data saved to {output_file}")

# Main function
def main():
    input_file = "servers.txt"
    output_file = "server_health_output.xlsx"

    username, password = read_credentials()
    clusters = read_servers(input_file)

    results = []
    with ThreadPoolExecutor(max_workers=10) as executor:
        for cluster, servers in clusters.items():
            results.extend(executor.map(lambda s: process_server(cluster, s, username, password), servers))

    save_to_excel(results, output_file)

if __name__ == "__main__":
    main()